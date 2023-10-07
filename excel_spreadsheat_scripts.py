from CardPool import CardPool
from Deck import Deck
from typing import List, Dict, Literal
import openpyxl
import xlsxwriter
import re

EXCELFILE: str = "created_deckroll_excel_spreadsheats/Deckroll-2023-09-12.xlsx"
NEW_EXCELFILE: str = "created_deckroll_excel_spreadsheats/Deckroll-2023-09-12-new.xlsx"
WORKSHEET_NAME: str = "rolled decks"
DECKLINK_TEXT: Literal["decklink", "regions", "champions"] = "regions"
DECKLINK_PREFIX: str = "https://runeterra.ar/decks/code/"

format: Literal["client_Formats_Eternal_name", "client_Formats_Standard_name"] = "client_Formats_Eternal_name"
card_pool = CardPool(format=format)

with xlsxwriter.Workbook(NEW_EXCELFILE) as new_workbook:
    new_worksheet = new_workbook.add_worksheet(WORKSHEET_NAME)
    old_workbook = openpyxl.load_workbook(EXCELFILE)
    old_worksheet = old_workbook[WORKSHEET_NAME]
    for row_index, row in enumerate(old_worksheet.iter_rows(values_only=True)):
        for col_index, value in enumerate(row):
            if isinstance(value, str):
                # if it's a link, only get the part after the last slash (for typical decklinks)
                if "/" in value:
                    value = value.rsplit('/', 1)[1]
                deckcode_pattern = r'^[A-Z0-9]+$'
                if re.match(deckcode_pattern, value):
                    deck = Deck(card_pool=card_pool)
                    deck.load_deck_from_deckcode(deckcode=value)

                    if DECKLINK_TEXT == "decklink":
                        new_worksheet.write(row_index, col_index, DECKLINK_PREFIX + deck.deckcode)
                    elif DECKLINK_TEXT == "regions":
                        new_worksheet.write_url(row=row_index, col=col_index, url=DECKLINK_PREFIX + deck.deckcode, string=" ".join(deck.regions))
                    elif DECKLINK_TEXT == "champions":
                        new_worksheet.write_url(row=row_index, col=col_index, url=DECKLINK_PREFIX + deck.deckcode, string=" ".join([champion.name for champion in deck.get_cards_by_card_type_sorted_by_cost_and_alphabetical("Champion")]))
                else:
                    new_worksheet.write(row_index, col_index, value)